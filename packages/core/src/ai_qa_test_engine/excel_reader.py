"""Excel data reader for test data injection.

Reads Excel files and makes cell data available as variables
for use in Gherkin steps via ${variable_name} substitution.

Usage in Gherkin:
    Given data from "TestData.xlsx" sheet "Login" row 1
    When I enter "${username}" in the username field
    And I enter "${password}" in the password field
"""

from pathlib import Path
from typing import Any

import openpyxl


def read_excel_data(
    file_path: Path,
    sheet_name: str,
    row: int | None = None,
    selection_set: int | None = None,
) -> dict[str, Any]:
    """Read data from an Excel file and return as a variables dictionary.

    Column headers (row 1) become variable names (lowercased, spaces→underscores).
    Data comes from the specified row.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the sheet to read
        row: Specific row number to read (1-indexed, excluding header)
        selection_set: Alternative to row — reads row N from data rows

    Returns:
        Dictionary mapping column_name → cell_value

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet or row not found
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {file_path.name}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # Read headers from row 1
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            # Normalize header: lowercase, spaces to underscores
            header = str(cell.value).strip().lower().replace(" ", "_")
            headers.append(header)
        else:
            headers.append(None)

    if not any(headers):
        raise ValueError(f"No headers found in row 1 of sheet '{sheet_name}'")

    # Determine which data row to read
    data_row_num = None
    if row is not None:
        data_row_num = row + 1  # +1 because row 1 is headers
    elif selection_set is not None:
        data_row_num = selection_set + 1  # selection_set is 1-indexed
    else:
        data_row_num = 2  # Default: first data row

    # Read the data row
    data_row = list(ws[data_row_num])
    if not data_row or all(cell.value is None for cell in data_row):
        raise ValueError(
            f"Row {data_row_num} is empty in sheet '{sheet_name}'"
        )

    # Build variables dict
    variables = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        if idx < len(data_row):
            value = data_row[idx].value
            variables[header] = value if value is not None else ""

    wb.close()
    return variables
